from unicodedata import name
from django.shortcuts import render

from apis.models import ApplicantExperienceModel, ApplicantInfoModel, ApplicantQualificationModel, CollegeInfoModel, EmployeeInfoModel, VacanciesInfoModel, VacancyApplicantMapping

from django.contrib.auth.models import User
from django.http import HttpResponse, JsonResponse
from django.contrib.auth import login, authenticate, logout
from rest_framework.views import APIView
from rest_framework.response import Response
import openpyxl
from rest_framework.decorators import api_view
# Create your views here.

class ApplicantAPIView(APIView):
    # Method to fetch the details of an applicant using username.
    def get(self, request, username, format = None):
        # if (pk == None):
        #     applicants = ApplicantInfoModel.objects.all()
        #     result = []
        #     for applicant in applicants:
        #         temp_result = {}
        #         temp_result["id"] = 
        #         temp_result["username"] = applicant.user.username
        #         temp_result["email"] = applicant.user.email
        #         temp_result["description"] = applicant.description
        #         temp_result["full_name"] = applicant.full_name
        #         temp_result["DOB"] = applicant.DOB
        #         temp_result["gender"] = applicant.gender
        #         temp_result["address"] = applicant.address
        #         temp_result["state"] = applicant.state
        #         temp_result["pincode"] = applicant.pincode
        #         temp_result["category"] = applicant.category
        #         temp_result["marital_status"] = applicant.marital_status
        #         temp_result["phone_number"] = applicant.phone_number
        #         temp_result["total_experience"] = applicant.total_experience
        #         temp_result["skillset"] = applicant.skillset
        #         result.append(temp_result)
        #     return Response(result, status = 200)

        # else:
        user = User.objects.get(username = username)
        applicant = ApplicantInfoModel.objects.get(user = user)
        temp_result = {}
        temp_result["username"] = applicant.user.username
        temp_result["email"] = applicant.user.email
        temp_result["description"] = applicant.description
        temp_result["full_name"] = applicant.full_name
        temp_result["DOB"] = applicant.DOB
        temp_result["gender"] = applicant.gender
        temp_result["address"] = applicant.address
        temp_result["state"] = applicant.state
        temp_result["pincode"] = applicant.pincode
        temp_result["category"] = applicant.category
        temp_result["marital_status"] = applicant.marital_status
        temp_result["phone_number"] = applicant.phone_number
        temp_result["total_experience"] = applicant.total_experience
        temp_result["skillset"] = applicant.skillset.split(" ")
        return Response(temp_result, status = 200)

    # Posting the data of the applicant and user signup and login api.
    def post(self, request, format = None):
        if (request.data["purpose"] == "signup"):
            username = request.data["username"]
            if (User.objects.filter(username = username).exists()):
                return Response({"mssg" : "username already exists!"}, status = 409)
            email = request.data["email"]
            if (User.objects.filter(email = email).exists()):
                return Response({"mssg" : "email already exists!"}, status = 409)
            password = request.data["password"]
            confirm_password = request.data["confirm_password"]
            if (password == confirm_password):
                user = User.objects.create(username = username, email = email, password = password)
                user.set_password(user.password)
                user.save()
                login(request, user)
                return Response({"mssg": "user signed up successfully!"}, status = 200)
            else:
                return Response({"mssg": "passwords do not match!"}, status = 409)
        elif (request.data["purpose"] == "login"):
            if (authenticate(username = request.data["username"], password = request.data["password"])):
                return Response({"mssg": "user logedin successfully!"}, status = 200)
            else:
                return Response({"mssg": "login failed!"}, status = 404)
        elif (request.data["purpose"] == "fill details"):
            user = User.objects.get(username = request.data["username"])
            request.data["details"]["user"] = user
            request.data["details"]["skillset"] = " ".join(request.data["details"]["skillset"])
            ApplicantInfoModel.objects.create(**request.data["details"])
            return Response({"mssg": "data updated successfully!"}, status = 202)

    # Updating the data of the applicant using username.
    def put(self, request, username, format = None):
        user = User.objects.get(username = username)
        applicant = ApplicantInfoModel.objects.get(user = user)
        user.email = request.data.get("email")
        applicant.description = request.data.get("description")
        applicant.full_name = request.data.get("full_name")
        applicant.DOB = request.data.get("DOB")
        applicant.gender = request.data.get("gender")
        applicant.address = request.data.get("address")
        applicant.state = request.data.get("state")
        applicant.pincode = request.data.get("pincode")
        applicant.category = request.data.get("category")
        applicant.marital_status = request.data.get("marital_status")
        applicant.phone_number = request.data.get("phone_number")
        applicant.total_experience = request.data.get("total_experience")
        applicant.skillset = " ".join(request.data.get("skillset"))
        user.save()
        applicant.save()
        return Response({"mssg" : "user updated successfully"}, status = 204)

    # Deleting user from the database.
    def delete(self, request, username, format = None):
        user = User.objects.get(username = username)
        applicant = ApplicantInfoModel.objects.get(user = user)
        user.delete()
        applicant.delete()
        return Response({"mssg": "user delete successfully"}, status = 200)



# Model for performing the CRUD operations on the user qualification.

class CollegeAPIView(APIView):
    def get(self, request, username, format = None):
        user = User.objects.get(username = username)
        college = CollegeInfoModel.objects.get(user = user)
        temp_result = {}
        temp_result["username"] = college.user.username
        temp_result["email"] = college.user.email
        temp_result["empid"] = college.empid
        temp_result["location"] = college.location
        temp_result["website"] = college.website
        temp_result["director_mail"] = college.director_mail
        temp_result["registrar_mail"] = college.registrar_mail
        temp_result["hod_mail"] = college.hod_mail
        return Response(temp_result, status = 200)  

    def post(self, request, format = None):
        if (request.data["purpose"] == "signup"):
            username = request.data["username"]
            if (User.objects.filter(username = username).exists()):
                return Response({"mssg" : "username already exists!"}, status = 409)
            email = request.data["email"]
            if (User.objects.filter(email = email).exists()):
                return Response({"mssg" : "email already exists!"}, status = 409)
            password = request.data["password"]
            confirm_password = request.data["confirm_password"]
            if (password == confirm_password):
                user = User.objects.create(username = username, email = email, password = password)
                user.set_password(user.password)
                user.save()
                login(request, user)
                return Response({"mssg": "user signed up successfully!"}, status = 200)
            else:
                return Response({"mssg": "passwords do not match!"}, status = 409)
        elif (request.data["purpose"] == "login"):
            if (authenticate(username = request.data["username"], password = request.data["password"])):
                return Response({"mssg": "user logedin successfully!"}, status = 200)
            else:
                return Response({"mssg": "login failed!"}, status = 404)
        elif (request.data["purpose"] == "fill details"):
            user = User.objects.get(username = request.data["username"])
            request.data["details"]["user"] = user
            CollegeInfoModel.objects.create(**request.data["details"])
            return Response({"mssg": "data updated successfully!"}, status = 202)  

    def put(self, request, username, format = None):
        user = User.objects.get(username = username)
        college = CollegeInfoModel.objects.get(user = user)
        user.email = request.data.get("email")
        college.empid = request.data.get("empid")
        college.location = request.data.get("location")
        college.website = request.data.get("website")
        college.director_mail = request.data.get("director_mail")
        college.registrar_mail = request.data.get("registrar_mail")
        college.hod_mail = request.data.get("hod_mail")
        user.save()
        college.save()
        return Response({"mssg" : "user updated successfully"}, status = 204)

    def delete(self, request, username, format = None):
        user = User.objects.get(username = username)
        college = CollegeInfoModel.objects.get(user = user)
        user.delete()
        college.delete()
        return Response({"mssg": "user delete successfully"}, status = 200)    

class QualificationAPIView(APIView):
    
    # Method to get the qualifications of a particular applicant using username.
    def get(self, request, username, format = None):
        applicant = User.objects.get(username = username)
        qualifications = ApplicantQualificationModel.objects.filter(applicant = applicant)
        result = []
        for qualification in qualifications:
            temp_result = {}
            temp_result["qualification_title"] = qualification.qualification_title
            temp_result["institute"] = qualification.institute
            temp_result["passing_year"] = qualification.passing_year
            temp_result["marks"] = qualification.marks
            result.append(temp_result)
        return Response(result, status = 200)

    # Posting the qualifications of an applicant using the username of the applicant.
    def post(self, request, username, format = None):
        user = User.objects.get(username = username)
        request.data["applicant"] = user
        ApplicantQualificationModel.objects.create(**request.data)
        return Response({"mssg" : "qualification added successfully"}, status = 202)

    # Method to update the qualifications of a user using the username and the qualification title.
    def put(self, request, username, format = None):
        applicant = User.objects.get(username = username)
        qualification = ApplicantQualificationModel.objects.get(applicant = applicant, qualification_title = request.data["title"])
        qualification.qualification_title = request.data["qualification_title"]
        qualification.institute = request.data["institute"]
        qualification.passing_year = request.data["passing_year"]
        qualification.marks = request.data["marks"]
        qualification.save()
        return Response({"mssg" : "qualification updated successfully"}, status = 204)

    # deleting the qualification of the user using username and the qualification title.
    def delete(self, request, username, format = None):
        applicant = User.objects.get(username = username)
        qualification = ApplicantQualificationModel.objects.get(applicant = applicant, qualification_title = request.data["qualification_title"])
        qualification.delete()
        return Response({"mssg": "qualification deleted successfully"}, status = 200)


# Method to get the employee data using the college name and the employee id. We need to mention get in the square brackets else nothing will work. 
@api_view(["GET"])
def get_employee_by_id(request, college_name, id):
    if (request.method == "GET"):
        user = User.objects.get(username = college_name)
        college = CollegeInfoModel.objects.get(user = user)
        employee = EmployeeInfoModel.objects.get(college = college, id = id)
        temp_result = employee.__dict__
        del temp_result["_state"]
        return Response({"employee" : temp_result}, status = 200)


# this api is used to change the status of the employee from active to notice period and mail all the required faculties that recruitment process has been initiated.
@api_view(["POST"])
def Change_employee_status(request, college_name, id):
    if (request.method == "POST"):
        user = User.objects.get(username = college_name)
        college = CollegeInfoModel.objects.get(user = user)
        employee = EmployeeInfoModel.objects.get(college = college, id = id)
        employee.status = request.data["status"]
        employee.save()
        return Response({"mssg": "status changed successfully!"}, status = 204)


    

class EmployeeAPIView(APIView):

    # Method to fetch the data of all the employees of the particular college using college name.
    def get(self, request, college_name, format = None):
        user = User.objects.get(username = college_name)
        college = CollegeInfoModel.objects.get(user = user)
        employees = EmployeeInfoModel.objects.filter(college = college)
        result = []
        for employee in employees:
            temp_result = {}
            temp = employee.__dict__
            print (temp)
            for key in temp:
                # This state is the reference object to the college.
                if (key == "_state"):
                    continue
                temp_result[key] = temp[key]
            result.append(temp_result)
        return Response({"employees" : result}, status = 200)

    # Method to post the data of the employees of a college using excel sheet or by using manual method using college name.
    def post(self, request, college_name, format = None):
        user = User.objects.get(username = college_name)
        college = CollegeInfoModel.objects.get(user = user)
        if (request.data["method"] == "excel file"):
            wb_obj = openpyxl.load_workbook(request.FILES["details"]) 
            sheet_obj = wb_obj.active 
            row = sheet_obj.max_row
            column = sheet_obj.max_column
            count = 0
            keys = ["college", "name", "DOB", "gender", "category", "status", "email", "phone_number", "designation"]
            for i in range(2, row + 1): 
                values = [college]
                for j in range(1, column + 1):
                    values.append(sheet_obj.cell(row = i, column = j).value)
                comb_list = zip(keys, values)
                data = dict(comb_list)
                EmployeeInfoModel.objects.create(**data)
                count += 1
            return Response({"mssg": "{} number of records created".format(count)}, status = 201)
        elif (request.data["method"] == "manual"):
            request.data["details"]["college"] = college
            EmployeeInfoModel.objects.create(**request.data["details"])
            return Response({"mssg": "employee added successfully!"}, status = 201)

    # Method to update the data of the employee using college name a.k.a. username and the employee id as there is no other unique parameter to be used.
    def put(self, request, college_name, format = None):
        user = User.objects.get(username = college_name)
        college = CollegeInfoModel.objects.get(user = user)
        employee = EmployeeInfoModel.objects.get(college = college, id = request.data["id"])
        employee.name = request.data["name"]
        employee.DOB = request.data["DOB"]
        employee.gender = request.data["gender"]
        employee.category = request.data["category"]
        employee.status = request.data["status"]
        employee.email = request.data["email"]
        employee.phone_number = request.data["phone_number"]
        employee.designation = request.data["designation"]
        employee.save()
        return Response({"mssg": "employee details updated successfully"}, status = 204)

    # Method to delete the record of an employee using college name and id of the employee.
    def delete(self, request, college_name, format = None):
        user = User.objects.get(username = college_name)
        college = CollegeInfoModel.objects.get(user = user)
        employee = EmployeeInfoModel.objects.get(college = college, id = request.data["id"])
        employee.delete()
        return Response({"mssg": "employee deleted successfully!"}, status = 200)


# Method to get the vacancies for a particular applicant using applicant username.
@api_view(["GET"])
def get_vacancies_for_applicant(request, username):
    user = User.objects.get(username = username)
    mappings = VacancyApplicantMapping.objects.filter(applicant = user)
    vacancies = []
    for mapping in mappings:
        vacancies.append(mapping.vacancy)
    result = []
    for vacancy in vacancies:
        temp_result = {}
        temp = vacancy.__dict__
        # print (temp)
        for key in temp:
            # This state is the reference object to the college.
            if (key == "_state"):
                continue
            if (key == "skills"):
                temp["skills"] = temp["skills"].split(" ")
            temp_result[key] = temp[key]
        temp_result["college_name"] = vacancy.college.user.username
        temp_result["location"] = vacancy.college.location
        temp_result["website"] = vacancy.college.website
        result.append(temp_result)
    return Response({"vacancies" : result}, status = 200)


# Method to get the applicants for a particular vacancy using vacancy id.
@api_view(["GET"])
def get_applicants_for_vacancy(request, id):
    mappings = VacancyApplicantMapping.objects.filter(vacancy = id)
    applicants = []
    for mapping in mappings:
        applicants.append(mapping.applicant)
        # print (mapping)
    result = []
    for applicant in applicants:
        temp_result = {}
        temp = applicant.__dict__
        # print (temp)
        for key in temp:
            if (key == "_state"):
                continue
            temp_result[key] = temp[key]
        result.append(temp_result)
    return Response({"applicants" : result}, status = 200)
            

class VacanciesAPIView(APIView):

    # Method to get all the vacancies posted by a particular college using college name or to get all the vacancies present in the system if college not given.
    def get(self, request, college_name = None, format = None):
        vacancies = VacanciesInfoModel.objects.all()
        if (college_name):
            user = User.objects.get(username = college_name)
            college = CollegeInfoModel.objects.get(user = user)
            vacancies = vacancies.filter(college = college)
        result = []
        for vacancy in vacancies:
            temp_result = {}
            temp = vacancy.__dict__
            print (temp)
            for key in temp:
                # This state is the reference object to the college.
                if (key == "_state"):
                    continue
                if (key == "skills"):
                    temp["skills"] = temp["skills"].split(" ")
                temp_result[key] = temp[key]
            result.append(temp_result)
        return Response({"vacancies" : result}, status = 200)

    # Method to post a new vacancy using college name or user name of the college.
    def post(self, request, college_name, format = None):
        user = User.objects.get(username = college_name)
        college = CollegeInfoModel.objects.get(user = user)
        request.data["college"] = college
        temp = " ".join(request.data["skills"])
        request.data["skills"] = temp
        VacanciesInfoModel.objects.create(**request.data)
        return Response({"mssg": "Vacancy posted successfully!"}, status = 201)

    # Method to update the details of a particular vacancy using college name and id of the vacancy.
    def put(self, request, college_name, format = None):
        user = User.objects.get(username = college_name)
        college = CollegeInfoModel.objects.get(user = user)
        vacancy = VacanciesInfoModel.objects.get(college = college, id = request.data["id"])
        vacancy.title = request.data["title"]
        vacancy.type = request.data["type"]
        vacancy.experience = request.data["experience"]
        vacancy.date_of_posting = request.data["date_of_posting"]
        vacancy.state = request.data["state"]
        vacancy.description = request.data["description"]
        vacancy.responsibilities = request.data["responsibilities"]
        vacancy.qualifications = request.data["qualifications"]
        vacancy.skills = " ".join(request.data["skills"])
        vacancy.compensation = request.data["compensation"]
        vacancy.save()
        return Response({"mssg": "vacancy details updated successfully!"}, status = 204)

    # Method to delete a particular vacancy using college name and id.
    def delete(self, request, college_name, format = None):
        user = User.objects.get(username = college_name)
        college = CollegeInfoModel.objects.get(user = user)
        vacancy = VacanciesInfoModel.objects.get(college = college, id = request.data["id"])
        vacancy.delete()
        return Response({"mssg": "Vacancy deleted successfully!"}, status = 200)
        
class ExperienceAPIView(APIView):
    
    def get(self, request, username, format = None):
        applicant = User.objects.get(username = username)
        experiences = ApplicantExperienceModel.objects.filter(applicant = applicant)
        result = []
        for experience in experiences:
            temp_result = {}
            temp_result["designation"] = experience.designation
            temp_result["from_date"] = experience.from_date
            temp_result["to_date"] = experience.to_date
            temp_result["institute"] = experience.institute
            temp_result["details"] = experience.details
            result.append(temp_result)
        return Response(result, status = 200)

    def post(self, request, username, format = None):
        user = User.objects.get(username = username)
        request.data["applicant"] = user
        ApplicantExperienceModel.objects.create(**request.data)
        return Response({"mssg" : "experience added successfully"}, status = 202)


    def put(self, request, username, format = None):
        applicant = User.objects.get(username = username)
        experience = ApplicantExperienceModel.objects.get(applicant = applicant, institute = request.data["institute"])
        experience.designation = request.data["designation"]
        experience.from_date = request.data["from_date"]
        experience.to_date = request.data["to_date"]
        experience.institute = request.data["institute"]
        experience.details = request.data["details"]
        experience.save()
        return Response({"mssg" : "experience updated successfully"}, status = 204)


    def delete(self, request, username, format = None):
        applicant = User.objects.get(username = username)
        experience = ApplicantExperienceModel.objects.get(applicant = applicant, institute = request.data["institute"])
        experience.delete()
        return Response({"mssg": "experience deleted successfully"}, status = 200)
